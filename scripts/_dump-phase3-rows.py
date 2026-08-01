import openpyxl, sys
from pathlib import Path

out = []
wb = openpyxl.load_workbook(Path(r"d:\cybsec-pmo\UAT_Test_Case_Register_v2_phase_3.xlsx"))
ws = wb["Test Cases"]

for r in range(2, ws.max_row + 1):
    phase = str(ws.cell(r, 2).value or "")
    if phase.strip() != "3 - Reporting":
        continue
    out.append("=" * 78)
    out.append(f"ROW {r} | {ws.cell(r, 1).value} | {ws.cell(r, 3).value} | {ws.cell(r, 7).value}")
    out.append(f"SCENARIO: {ws.cell(r, 9).value}")
    out.append(f"STEPS:\n{ws.cell(r, 11).value}")
    out.append(f"EXPECTED_NOW: {ws.cell(r, 12).value}")
    out.append(f"EVIDENCE: {ws.cell(r, 13).value}")

out.append(f"\nTOTAL ROWS: {sum(1 for line in out if line.startswith('ROW '))}")
Path(r"d:\cybsec-pmo\scripts\_phase3-dump.txt").write_text("\n".join(out), encoding="utf-8")
print("written", len(out), "lines")
