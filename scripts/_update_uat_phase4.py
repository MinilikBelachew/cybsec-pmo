"""Update Phase 4 Risk & Compliance UAT rows 107-136 (Test Steps + Expected Result)."""
from pathlib import Path

import openpyxl

PATH = Path(r"d:\cybsec-pmo\UAT_Test_Case_Register_v2_phase_3.xlsx")

# Excel row -> (Test Steps, Expected Result)
UPDATES: dict[int, tuple[str, str]] = {
    107: (
        """1. Log in as a PM or PMO Lead with risks:edit.
2. Navigate to Risk & Issues > Risk Register (/dashboard/risks).
3. Click Add risk. Select a project, enter Title, Category (e.g. TECHNICAL), Impact (1–4), Likelihood (1–4), Owner, Mitigation plan, Target date, optional Residual impact/likelihood, and Status (Open).
4. Save. Confirm the risk appears in the register with Score = Impact × Likelihood and residual rating when residual values are provided.
5. Open a project workspace > Risks tab and confirm the same risk is listed for that project.""",
        """Risk is created with category, impact, likelihood, auto score (impact × likelihood), owner, mitigation, target date, residual rating (when residual values are set), and status. It appears on Risk Register and on the project workspace Risks tab. Create returns HTTP 201/200 with no validation errors.""",
    ),
    108: (
        """1. Log in as a PM/PMO Lead with risks:edit and open Risk & Issues > Risk Register.
2. Click Add risk and attempt to save with Title blank, Owner unset, or Impact/Likelihood outside 1–4.
3. Confirm client-side validation messages block submit.
4. Optionally call POST /v1/projects/{projectId}/risks with missing required fields and confirm HTTP 400.""",
        """Required fields (title, category, impact, likelihood, owner) are enforced. Invalid or incomplete submissions are rejected with clear validation messages (UI and/or HTTP 400). No incomplete risk row is persisted.""",
    ),
    109: (
        """1. Log in as an Engineer without risks:edit (view-only or no risks permission as seeded).
2. Open /dashboard/risks. Confirm Add risk is hidden or the page shows a permission message.
3. Attempt POST /v1/projects/{projectId}/risks with a valid body and confirm HTTP 403.
4. Log in as PM/PMO Lead with risks:edit and confirm create/update/close are available.""",
        """Permissions are enforced. Users without risks:edit cannot create or mutate risks (UI gated and API returns 403). Users with risks:edit can manage risks within their project scope. No unauthorized write occurs.""",
    ),
    110: (
        """1. Log in as a PM/PMO Lead and open Risk Register > Add risk.
2. Set Impact=3 and Likelihood=4; leave residual blank; save.
3. Confirm Score displays as 12 on the new row.
4. Edit the risk to Impact=2, Likelihood=2; save and confirm Score updates to 4.
5. Set Residual impact=1 and Residual likelihood=2; save and confirm residual rating = 2.""",
        """Score is auto-calculated as Impact × Likelihood (integer 1–4 matrix) on create and update. Residual rating is Impact_residual × Likelihood_residual when residual values are provided. UI and API score fields match the formula.""",
    ),
    111: (
        """1. Ensure an Alert Catalogue rule exists for eventType RISK_SCORE_BREACHED with thresholdConfig.scoreGte=12 (or create one under Alert Catalogue).
2. As PM/PMO Lead, create or update a risk so Score ≥ 12 (e.g. Impact=4, Likelihood=3).
3. Confirm owner receives a High risk score / RISK_THRESHOLD_BREACHED notification and an AlertEvent appears under Alert Catalogue > Instances (deliveryStatus queued/sent).
4. Open Dashboard (/dashboard) as PMO Lead and confirm Active risks / risk KPI reflects open Risk register rows (not merely high-priority tasks).
5. Confirm the dashboard risk matrix (if shown) includes the live risk.""",
        """When score reaches/crosses the high threshold (default 12), threshold notification and matching alert-catalogue event fire. Dashboard active-risk KPI counts open Risk table rows (status not Closed/Cancelled) for scoped projects. Risk matrix/KPI linkage uses live risk data.""",
    ),
    112: (
        """1. Log in as a PM/PMO Lead with issues:edit.
2. Navigate to Risk & Issues > Issue Tracker (/dashboard/issues).
3. Click Add issue. Select project, Title, Priority (Low/Medium/High/Critical), Owner, Due date, optional Expected resolution date, Status Open.
4. Save. Confirm the issue appears with those fields.
5. Close the issue with a resolution note (and evidence key if used) and confirm Status=Closed and resolution/evidence are stored.
6. Open project workspace > Issues tab and confirm the issue is listed.""",
        """Issue is created with priority, owner, due date, status, and optional expected resolution date. Closure captures resolution note (and evidence when provided). Issue appears on Issue Tracker and project workspace Issues tab. APIs return success without validation errors.""",
    ),
    113: (
        """1. As PM/PMO Lead open Issue Tracker > Add issue.
2. Set Expected resolution date to a future date and save.
3. Confirm Expected resolution date is shown on the issue row/detail.
4. Edit the date and confirm it updates; clear it if the UI allows and confirm null is accepted.""",
        """Expected resolution date is captured on create/update, displayed on the issue, and persisted via the issues API. Changing or clearing the date updates the stored value accordingly.""",
    ),
    114: (
        """1. Ensure an Alert Catalogue rule exists for eventType ISSUE_ESCALATED (optional threshold none).
2. Create an issue with Priority=High or Critical, or with Due date in the past and Status Open.
3. Confirm ISSUE_ESCALATED notification is sent to the owner (and alert engine fires an AlertEvent when a catalogue rule matches).
4. Confirm requiresEscalation / overdue indicators appear on the issue in the UI.""",
        """Overdue and/or High/Critical open issues trigger escalation notification (ISSUE_ESCALATED). When a matching alert rule exists, an AlertEvent is created. UI shows overdue/escalation indicators for those issues.""",
    ),
    115: (
        """1. As PM/PMO Lead, open an Open issue on Issue Tracker.
2. Close the issue with a resolution note.
3. Confirm raiser and owner receive ISSUE closed / closure notification.
4. Confirm Status=Closed and closed/updated timestamps are set.""",
        """Closing an issue notifies relevant parties (raiser/owner). Status becomes Closed with resolution note retained. Closure notification is visible in Notifications.""",
    ),
    116: (
        """1. Log in as PMO Lead / admin with notifications:manage.
2. Open Risk & Issues > Alert Catalogue (/dashboard/alerts).
3. Click Add rule. Set Event type=RISK_SCORE_BREACHED and Score threshold (≥)=12 (thresholdConfig.scoreGte).
4. Save. Confirm the rule appears in Catalogue rules with the threshold JSON shown.
5. Create a risk with score below threshold and confirm no new AlertEvent; then raise score ≥ 12 and confirm an event fires.""",
        """Alert thresholds are configurable per catalogue rule (e.g. scoreGte). Rules persist and are listed on Alert Catalogue. Events fire only when the metric meets the configured threshold.""",
    ),
    117: (
        """1. As PMO Lead open Alert Catalogue > Add rule.
2. Set Channels to in_app,email (comma-separated) and save.
3. Confirm the Channels column shows in_app, email.
4. Trigger a matching alert (e.g. high risk score) and confirm AlertEvent rows are created per channel with deliveryStatus queued/sent.""",
        """Channels are configurable on each rule (in_app and/or email). Saved channels display on the catalogue table. Fired alerts create per-channel AlertEvent instances with delivery status tracked.""",
    ),
    118: (
        """1. As PMO Lead open Alert Catalogue > Add rule.
2. Under Recipient roles, select one or more roles (e.g. PMO Lead, PM) via the checkboxes.
3. Save. Confirm the Recipients column lists the selected role labels.
4. Fire a matching alert and confirm users with those roles receive the notification.""",
        """Recipient roles are configurable on create and shown on the catalogue. When an alert fires, active users in the selected roles receive notifications. Empty recipient set is allowed but delivers to no role-resolved users.""",
    ),
    119: (
        """1. As PMO Lead open Alert Catalogue > Add rule.
2. Set Reminder cadence (hrs) to a low test value (e.g. 1) and save.
3. Fire an alert and leave it unacknowledged.
4. Wait for / trigger the alert reminder job (ALERT_REMINDER_CRON, default every 15 minutes) and confirm a reminder notification is sent and nextReminderAt advances.""",
        """Reminder cadence (hours) is stored on the rule. Unacknowledged AlertEvents receive reminder notifications on the configured cadence via the alert reminder scheduler. nextReminderAt updates after each reminder.""",
    ),
    120: (
        """1. As a user with notifications:view, open Alert Catalogue and locate Instances.
2. Identify an unacknowledged AlertEvent (ackedAt empty).
3. Click Acknowledge.
4. Confirm ackedAt/acknowledgedBy are set and further reminders/escalation stop for that event.""",
        """Acknowledgement is supported on alert instances. After ack, the event shows acknowledged metadata and reminder/escalation progression for that event stops.""",
    ),
    121: (
        """1. As PMO Lead open Alert Catalogue > Add rule.
2. Set Escalation role (e.g. pmo_lead) and Escalation delay (hrs) to a low test value.
3. Save. Confirm Escalation column shows role / delay.
4. Fire an alert, leave it unacked past the delay (or adjust timestamps / run escalation path in the engine), and confirm escalation notification to the escalation role and escalationLevel increments.""",
        """Escalation hierarchy is defined per rule via escalationRole and escalationDelayHrs. Unacknowledged alerts past the delay notify the escalation role and increment escalationLevel. Catalogue UI shows the configured hierarchy.""",
    ),
    122: (
        """1. Create/activate a catalogue rule and force a failed delivery (or mark an AlertEvent deliveryStatus=failed/retrying with nextReminderAt due).
2. Wait for / run the alert retry job (ALERT_RETRY_CRON, default every 10 minutes).
3. Confirm retries follow backoff 1h → 4h → 12h (max 3 attempts) and status moves retrying → sent on success or failed after exhaustion.
4. Confirm docs/alert-catalogue.md documents this retry behaviour.""",
        """Retry behaviour is defined: failed deliveries enter retrying with backoff 1h, 4h, 12h (max 3 attempts), then failed. The retry scheduler processes due retries. Behaviour matches the alert catalogue documentation.""",
    ),
    123: (
        """1. Open repository doc docs/alert-catalogue.md (or the UAT-published catalogue artefact).
2. Confirm it lists event types (e.g. RISK_SCORE_BREACHED, ISSUE_ESCALATED), channels, recipients/hierarchy, reminder cadence, retry backoff, and acknowledgement.
3. Cross-check one live rule on /dashboard/alerts against the documented fields.
4. Note Cybsec business approval of the catalogue as a sign-off step for Gate 4 evidence.""",
        """Alert catalogue is documented (event types, thresholds, channels, recipients, cadence, escalation, retry, ack). Documented fields align with the Alert Catalogue UI/API. Formal Cybsec approval of the catalogue is recorded as Gate 4 evidence when provided.""",
    ),
    124: (
        """1. Log in as a PM/PMO Lead and open Risk & Issues > Escalations (/dashboard/escalations).
2. Create an escalation: select Project, Customer, Severity (e.g. High), SLA target hours, Owner, optional initial communication.
3. Save. Confirm the record shows customer, severity, SLA hours, owner, status Open, and communication if provided.
4. Close with a resolution summary and confirm closure fields update.""",
        """Customer escalation captures customer, severity, SLA target hours, owner, communication, resolution, and closure. Create opens status Open; close stores resolution summary and closedAt. Record is visible on Escalations list.""",
    ),
    125: (
        """1. Open an existing Open escalation on /dashboard/escalations.
2. Add a communication with Channel (e.g. Email) and content.
3. Confirm the communication appears in the escalation’s communication history with logger and timestamp.""",
        """Customer communications are recorded against the escalation with channel, content, logger, and timestamp. History remains visible on the escalation detail/list after save.""",
    ),
    126: (
        """1. Create an escalation with a short SLA (e.g. 1 hour) or High/Critical severity.
2. Confirm High/Critical notify management on create (ESCALATION_MANAGEMENT).
3. For SLA breach: leave Open past slaTargetHrs (or adjust createdAt) and wait for / run Escalation SLA job (ESCALATION_SLA_CRON, default every 15 minutes).
4. Confirm slaBreached=true and management notification is sent.""",
        """High/Critical escalations notify management on create. Open escalations past SLA are marked slaBreached by the SLA scheduler and escalate to management. Notifications are visible to management roles / owner.""",
    ),
    127: (
        """1. Open an Open escalation and choose Close.
2. Enter Resolution summary and confirm.
3. Confirm Status=Closed, closedAt set, and slaBreached reflects whether SLA was exceeded at close.
4. Confirm closure notification (ESCALATION_CLOSED) is received by the owner.""",
        """Closure is tracked with status Closed, resolution summary, closedAt, and slaBreached flag. Owner receives closure notification. Closed escalations no longer appear as open in filters.""",
    ),
    128: (
        """1. Log in as PM/Team Lead (action-point manager role) and open a project workspace > Action points tab.
2. Click Add action point. Create one with Source=Project (default).
3. Create another with Source=Task and select a linked task; then Source=Risk and Source=Issue with linked entities.
4. Confirm each row shows the Source badge (Project/Task/Risk/Issue).
5. Optionally create Meeting/MoM-linked actions from meeting flows (backend supports Meeting/MoM sourceType) and confirm they appear on Action Points portfolio (/dashboard/actions).""",
        """Action points can link to Project, Task, Risk, and Issue from the workspace Action points UI (source picker + linked entity). Meeting/MoM sources are supported by the API/meeting flows and appear in the portfolio. Source type is shown on each action row.""",
    ),
    129: (
        """1. In project workspace > Action points, click Add action point.
2. Enter Name, Owner (from project assignees), Due date (within project start/end), Priority (Low/Medium/High/Critical), leave Status Open.
3. Save. Confirm the list shows owner, due date, priority, and status.
4. Update status to In Progress then Done; as assignee-only user confirm Cancelled is not offered.""",
        """Action points require/store owner, due date, priority, and status. Due date is constrained to the project date range. Managers can set all statuses including Cancelled; assignees can update Open/In Progress/Done only. Values persist after refresh.""",
    ),
    130: (
        """1. Create an open action point with Due date within the next 3 days.
2. As a manager on /dashboard/actions, click Send due reminders and confirm toast shows sent count.
3. Confirm ACTION_POINT_REMINDER notification reaches the owner.
4. Optionally wait for / run the scheduled reminder job (ACTION_POINT_REMINDER_CRON, default daily 08:00) and confirm reminders are also sent automatically.""",
        """Reminders are sent for open action points due within 3 days via manual Send due reminders and/or the daily reminder scheduler. Owners receive ACTION_POINT_REMINDER notifications. Sent count is reported in the UI toast.""",
    ),
    131: (
        """1. Create an open action point with Due date in the past (or adjust due_date in DB for UAT).
2. Confirm the row shows Overdue in the Action points panel / portfolio.
3. Wait for / run the overdue job (ACTION_POINT_OVERDUE_CRON, default daily 09:00), or create/update an already-overdue action to trigger notify-on-write.
4. Confirm ACTION_POINT_OVERDUE notification is sent to the owner.""",
        """Overdue open action points are flagged in the UI. Overdue escalation/notification is sent to the owner via write-time notify and/or the dedicated overdue scheduler. Notification type ACTION_POINT_OVERDUE is recorded.""",
    ),
    132: (
        """1. Log in as PM/PMO Lead and open Risk & Issues > Action Points (/dashboard/actions).
2. Confirm closure report KPIs: Total, Closed, Overdue open.
3. Confirm breakdowns by owner and by status (and source type if shown).
4. Close an action point (status Done with optional closure note) and refresh; confirm Closed count increases and the item leaves overdue-open if applicable.""",
        """Closure reporting is available on the Action Points portfolio: totals for total/closed/overdue-open plus breakdowns by owner and status. Closing actions updates the report figures after refresh.""",
    ),
    133: (
        """1. Log in as a user with projects:edit and open Risk & Issues > Lessons Learned (/dashboard/lessons).
2. Click Add lesson. Select Category (e.g. DEPLOYMENT), enter Description and Recommendation, optional Tags and Project.
3. Save. Confirm the lesson appears in the list with category, description, recommendation, and author.""",
        """Lessons are captured with category, description/context, recommendation, optional tags/project, and author. The new lesson appears on Lessons Learned after save (HTTP 201/200).""",
    ),
    134: (
        """1. Open Lessons Learned with several lessons present (different categories/tags).
2. Filter by Category and confirm the list narrows.
3. Use search (q) for a word in description/recommendation and confirm matching lessons only.
4. Optionally filter by tag/project if exposed and confirm results.""",
        """Lessons are searchable/filterable by category, free-text query (description/recommendation/category), and tag/project where provided. Filters narrow the list without errors; clearing filters restores the full set.""",
    ),
    135: (
        """1. Ensure at least one lesson exists for a department/category used by a test project.
2. As PM/PMO Lead, open Projects > Create Project. Confirm Surfaced lessons panel (“Lessons for project setup”) loads relevant lessons.
3. Edit an existing project and set Status to Pending Closure. Confirm “Lessons for project closure” panel surfaces related lessons.
4. Confirm GET /v1/lessons/surface returns the same set used by the UI.""",
        """Relevant lessons are surfaced during project setup (create project sheet) and project closure (Pending Closure). The surfaced panel lists category, description, and recommendation from /lessons/surface. Empty state shows when no matches exist.""",
    ),
    136: (
        """1. Log in as an Engineer without projects:edit (view-only as seeded).
2. Open /dashboard/lessons. Confirm Add lesson is hidden or the page denies create.
3. Attempt POST /v1/lessons with a valid body and confirm HTTP 403.
4. Log in as PM/PMO Lead with projects:edit and confirm create/update succeed.
5. Confirm lessons APIs are gated by projects:view (read) and projects:edit (write) — dedicated lessons module permission is not required in current build.""",
        """Permission controls are enforced via projects:view/edit on lessons endpoints and UI. Unauthorized create/update returns 403 and does not persist. Authorized project editors can capture and update lessons. (Dedicated lessons module permission is out of scope for current Gate 4 build; projects permissions apply.)""",
    ),
}


def main() -> None:
    wb = openpyxl.load_workbook(PATH)
    ws = wb["Test Cases"]
    for row, (steps, expected) in UPDATES.items():
        tc = ws.cell(row, 1).value
        phase = ws.cell(row, 2).value
        if "4" not in str(phase):
            raise SystemExit(f"Row {row} is not Phase 4: {phase}")
        ws.cell(row, 11).value = steps.strip()
        ws.cell(row, 12).value = expected.strip()
        print(f"Updated {row} {tc}")
    wb.save(PATH)
    print(f"Saved {PATH}")


if __name__ == "__main__":
    main()
