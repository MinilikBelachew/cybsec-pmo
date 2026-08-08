"""
Generate stress-test fixtures: 2 projects × ~1000 tasks (+ phases + milestones).

Outputs (same folder):
  - load-test-2x1000-portfolio.xml   (MSPDI — Import MPP)
  - load-test-2x1000-portfolio.xlsx  (Projects + Phases/Tasks/Milestones sheets)

Run:
  python scripts/fixtures/load-test-2x1000/gen-load-test-2x1000.py
"""

from __future__ import annotations

import html
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

OUT_DIR = Path(__file__).resolve().parent

PROJECTS = [
    {
        "name": "LoadTest Alpha",
        "objective": "Stress-test import — Project A (1000 tasks).",
        "department": "Application Security",
        "customer": "Acme Financial Services",
    },
    {
        "name": "LoadTest Beta",
        "objective": "Stress-test import — Project B (1000 tasks).",
        "department": "Cloud Security",
        "customer": "Globex Manufacturing",
    },
]

PHASE_COUNT = 10
TASKS_PER_PHASE = 100  # 10 × 100 = 1000 leaf tasks per project
MILESTONES_PER_PROJECT = PHASE_COUNT  # one milestone per phase
START = date(2026, 1, 5)


def add_days(d: date, days: int) -> date:
    return d + timedelta(days=days)


def iso_dt(d: date, end: bool = False) -> str:
    return f"{d.isoformat()}T{'17' if end else '08'}:00:00"


def phase_window(phase_index: int) -> tuple[date, date]:
    # ~20 calendar days per phase, sequential
    start = add_days(START, phase_index * 20)
    end = add_days(start, 19)
    return start, end


def task_window(phase_index: int, task_index: int) -> tuple[date, date]:
    phase_start, _ = phase_window(phase_index)
    # Spread tasks across the phase (2 days each, wrapping)
    start = add_days(phase_start, (task_index % 10) * 2)
    end = add_days(start, 1)
    return start, end


def escape(text: str) -> str:
    return html.escape(text, quote=False)


def task_xml(
    *,
    uid: int,
    tid: int,
    name: str,
    outline_level: int,
    outline_number: str,
    wbs: str,
    summary: bool,
    start: date,
    finish: date,
    duration: str,
    milestone: bool = False,
    predecessors: list[int] | None = None,
) -> str:
    preds = predecessors or []
    pred_xml = "\n".join(
        f"""      <PredecessorLink>
        <PredecessorUID>{p}</PredecessorUID>
        <Type>1</Type>
        <CrossProject>0</CrossProject>
        <LinkLag>0</LinkLag>
        <LagFormat>7</LagFormat>
      </PredecessorLink>"""
        for p in preds
    )
    return f"""    <Task>
      <UID>{uid}</UID>
      <ID>{tid}</ID>
      <Name>{escape(name)}</Name>
      <Type>{1 if summary else 0}</Type>
      <IsNull>0</IsNull>
      <OutlineLevel>{outline_level}</OutlineLevel>
      <OutlineNumber>{outline_number}</OutlineNumber>
      <WBS>{wbs}</WBS>
      <Summary>{1 if summary else 0}</Summary>
      <Critical>{0 if summary else 1}</Critical>
      <Priority>500</Priority>
      <Start>{iso_dt(start)}</Start>
      <Finish>{iso_dt(finish, end=True)}</Finish>
      <Duration>{duration}</Duration>
      <Manual>0</Manual>
      <PercentComplete>0</PercentComplete>
      <Milestone>{1 if milestone else 0}</Milestone>
{pred_xml}
    </Task>"""


def build_xml() -> tuple[str, dict]:
    uid = 1
    tid = 1
    tasks: list[str] = []
    stats = {"projects": 0, "phases": 0, "leaf_tasks": 0, "milestones": 0, "total_rows": 0}

    portfolio_end = add_days(START, PHASE_COUNT * 20)

    tasks.append(
        task_xml(
            uid=0,
            tid=0,
            name="LoadTest Portfolio 2x1000",
            outline_level=0,
            outline_number="0",
            wbs="0",
            summary=True,
            start=START,
            finish=portfolio_end,
            duration="PT1600H0M0S",
        )
    )

    for p_idx, project in enumerate(PROJECTS, start=1):
        stats["projects"] += 1
        project_uid = uid
        uid += 1
        tid += 1
        proj_end = add_days(START, PHASE_COUNT * 20)
        tasks.append(
            task_xml(
                uid=project_uid,
                tid=tid - 1,
                name=project["name"],
                outline_level=1,
                outline_number=str(p_idx),
                wbs=str(p_idx),
                summary=True,
                start=START,
                finish=proj_end,
                duration="PT1600H0M0S",
            )
        )

        last_leaf_uid: int | None = None
        for ph in range(PHASE_COUNT):
            stats["phases"] += 1
            phase_no = ph + 1
            phase_start, phase_end = phase_window(ph)
            phase_uid = uid
            uid += 1
            tid += 1
            phase_name = f"Phase {phase_no:02d} — Workstream {phase_no}"
            tasks.append(
                task_xml(
                    uid=phase_uid,
                    tid=tid - 1,
                    name=phase_name,
                    outline_level=2,
                    outline_number=f"{p_idx}.{phase_no}",
                    wbs=f"{p_idx}.{phase_no}",
                    summary=True,
                    start=phase_start,
                    finish=phase_end,
                    duration="PT160H0M0S",
                )
            )

            for t in range(TASKS_PER_PHASE):
                stats["leaf_tasks"] += 1
                task_no = t + 1
                t_start, t_end = task_window(ph, t)
                is_milestone = task_no == TASKS_PER_PHASE
                if is_milestone:
                    stats["milestones"] += 1
                    name = f"Milestone {phase_no:02d} — Phase {phase_no} Complete"
                    duration = "PT0H0M0S"
                else:
                    name = f"Task {p_idx}.{phase_no}.{task_no:03d}"
                    duration = "PT16H0M0S"

                task_uid = uid
                uid += 1
                tid += 1
                preds: list[int] = []
                # Light dependency chain every 10th task (keeps file realistic, not fully serial)
                if task_no > 1 and task_no % 10 == 1 and last_leaf_uid is not None:
                    preds = [last_leaf_uid]

                tasks.append(
                    task_xml(
                        uid=task_uid,
                        tid=tid - 1,
                        name=name,
                        outline_level=3,
                        outline_number=f"{p_idx}.{phase_no}.{task_no}",
                        wbs=f"{p_idx}.{phase_no}.{task_no}",
                        summary=False,
                        start=t_start if not is_milestone else phase_end,
                        finish=t_end if not is_milestone else phase_end,
                        duration=duration,
                        milestone=is_milestone,
                        predecessors=preds,
                    )
                )
                last_leaf_uid = task_uid

    stats["total_rows"] = len(tasks)

    xml = f"""<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<!--
  Cybsec PMO load-test fixture — MSPDI portfolio.
  Import via Projects → Import MPP (accepts .xml).

  Shape:
  L0 = portfolio root
  L1 = 2 projects ({PROJECTS[0]['name']}, {PROJECTS[1]['name']})
  L2 = {PHASE_COUNT} phases per project
  L3 = {TASKS_PER_PHASE} tasks per phase ({TASKS_PER_PHASE * PHASE_COUNT} leaf tasks / project;
       last task in each phase is a milestone)

  Generated by gen-load-test-2x1000.py — do not hand-edit.
-->
<Project xmlns="http://schemas.microsoft.com/project">
  <Name>LoadTest Portfolio 2x1000</Name>
  <Title>LoadTest Portfolio 2x1000</Title>
  <Company>CyberSec IT</Company>
  <Author>Cybsec PMO Load Fixture</Author>
  <CreationDate>2026-01-01T09:00:00</CreationDate>
  <LastSaved>2026-01-01T09:00:00</LastSaved>
  <ScheduleFromStart>1</ScheduleFromStart>
  <StartDate>{iso_dt(START)}</StartDate>
  <FinishDate>{iso_dt(portfolio_end, end=True)}</FinishDate>
  <CalendarUID>1</CalendarUID>
  <DefaultStartTime>08:00:00</DefaultStartTime>
  <DefaultFinishTime>17:00:00</DefaultFinishTime>
  <MinutesPerDay>480</MinutesPerDay>
  <MinutesPerWeek>2400</MinutesPerWeek>
  <DaysPerMonth>20</DaysPerMonth>
  <CurrencyCode>USD</CurrencyCode>
  <Calendars>
    <Calendar>
      <UID>1</UID>
      <Name>Standard</Name>
      <IsBaseCalendar>1</IsBaseCalendar>
      <BaseCalendarUID>-1</BaseCalendarUID>
      <WeekDays>
        <WeekDay><DayType>1</DayType><DayWorking>0</DayWorking></WeekDay>
        <WeekDay><DayType>2</DayType><DayWorking>1</DayWorking><WorkingTimes><WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime><WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime></WorkingTimes></WeekDay>
        <WeekDay><DayType>3</DayType><DayWorking>1</DayWorking><WorkingTimes><WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime><WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime></WorkingTimes></WeekDay>
        <WeekDay><DayType>4</DayType><DayWorking>1</DayWorking><WorkingTimes><WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime><WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime></WorkingTimes></WeekDay>
        <WeekDay><DayType>5</DayType><DayWorking>1</DayWorking><WorkingTimes><WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime><WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime></WorkingTimes></WeekDay>
        <WeekDay><DayType>6</DayType><DayWorking>1</DayWorking><WorkingTimes><WorkingTime><FromTime>08:00:00</FromTime><ToTime>12:00:00</ToTime></WorkingTime><WorkingTime><FromTime>13:00:00</FromTime><ToTime>17:00:00</ToTime></WorkingTime></WorkingTimes></WeekDay>
        <WeekDay><DayType>7</DayType><DayWorking>0</DayWorking></WeekDay>
      </WeekDays>
    </Calendar>
  </Calendars>
  <Tasks>
{chr(10).join(tasks)}
  </Tasks>
</Project>
"""
    return xml, stats


def sheet_name(project_name: str, suffix: str) -> str:
    # Excel sheet names max 31 chars
    name = f"{project_name} {suffix}"
    return name[:31]


def build_xlsx() -> Path:
    wb = Workbook()
    ws_projects = wb.active
    ws_projects.title = "Projects"
    ws_projects.append(
        [
            "Name",
            "Objective",
            "Department",
            "Customer",
            "Engagement Type",
            "Billing Model",
            "Priority",
            "Start Date",
            "End Date",
            "Value",
            "Currency",
            "Primary PM",
            "Secondary PM",
            "Status",
        ]
    )

    portfolio_end = add_days(START, PHASE_COUNT * 20)

    for i, project in enumerate(PROJECTS):
        ws_projects.append(
            [
                project["name"],
                project["objective"],
                project["department"],
                project["customer"],
                "ManagedServices",
                "TimeAndMaterial",
                "High",
                START.isoformat(),
                portfolio_end.isoformat(),
                250000 + i * 50000,
                "USD",
                "John Smith",
                None,
                "Draft",
            ]
        )

        # Phases sheet
        ws_phases = wb.create_sheet(sheet_name(project["name"], "Phases"))
        ws_phases.append(
            ["Name", "Description", "Order", "Status", "Start Date", "End Date"]
        )
        phase_names: list[str] = []
        for ph in range(PHASE_COUNT):
            phase_no = ph + 1
            phase_start, phase_end = phase_window(ph)
            phase_name = f"Phase {phase_no:02d} — Workstream {phase_no}"
            phase_names.append(phase_name)
            ws_phases.append(
                [
                    phase_name,
                    f"Load-test phase {phase_no} for {project['name']}",
                    phase_no,
                    "Planned",
                    phase_start.isoformat(),
                    phase_end.isoformat(),
                ]
            )

        # Tasks sheet
        ws_tasks = wb.create_sheet(sheet_name(project["name"], "Tasks"))
        ws_tasks.append(
            [
                "Title",
                "Description",
                "Priority",
                "Status",
                "Assignee",
                "Phase",
                "Start Date",
                "End Date",
                "Effort Hours",
            ]
        )
        for ph in range(PHASE_COUNT):
            phase_name = phase_names[ph]
            for t in range(TASKS_PER_PHASE):
                task_no = t + 1
                t_start, t_end = task_window(ph, t)
                # Keep last task of each phase as a normal task in Tasks sheet;
                # milestones live on the Milestones sheet (matches UAT workbook shape).
                ws_tasks.append(
                    [
                        f"Task {ph + 1}.{task_no:03d}",
                        f"Generated load-test task {task_no} in {phase_name}",
                        "Medium" if task_no % 5 else "High",
                        "To_Do",
                        None,
                        phase_name,
                        t_start.isoformat(),
                        t_end.isoformat(),
                        16,
                    ]
                )

        # Milestones sheet (weights sum to 100%)
        ws_ms = wb.create_sheet(sheet_name(project["name"], "Milestones"))
        ws_ms.append(["Title", "Target Date", "Weight (%)", "Status", "Phase"])
        weight = 100 // MILESTONES_PER_PROJECT
        remainder = 100 - weight * MILESTONES_PER_PROJECT
        for ph in range(PHASE_COUNT):
            _, phase_end = phase_window(ph)
            w = weight + (remainder if ph == PHASE_COUNT - 1 else 0)
            ws_ms.append(
                [
                    f"Milestone {ph + 1:02d} — Phase {ph + 1} Complete",
                    phase_end.isoformat(),
                    w,
                    "Pending",
                    phase_names[ph],
                ]
            )

    out = OUT_DIR / "load-test-2x1000-portfolio.xlsx"
    wb.save(out)
    return out


def main() -> None:
    xml, stats = build_xml()
    xml_path = OUT_DIR / "load-test-2x1000-portfolio.xml"
    xml_path.write_text(xml, encoding="utf-8")

    xlsx_path = build_xlsx()

    print("wrote", xml_path)
    print("wrote", xlsx_path)
    print("stats", stats)
    print(
        f"xlsx: 2 projects × {PHASE_COUNT} phases × {TASKS_PER_PHASE} tasks "
        f"+ {MILESTONES_PER_PROJECT} milestones each"
    )


if __name__ == "__main__":
    main()
