"""
Workspace Import Tasks fixture for "DLP & DC Project Plan".

Phases + dates match the live project (imported from descount.mpp):
  Phase 1 - Project Initiation & Planning  2026-04-01 .. 2026-04-02
  Phase 2- Design & Documentation          2026-04-05 .. 2026-04-29
  Phase 3- Implementation & UAT            2026-06-29 .. 2026-09-13
  Phase 4- Project Closure                 2026-09-13 .. 2026-09-16

Outputs:
  - workspace-tasks-20000.xlsx   (sheet: Tasks)
  - workspace-phases-20000.xlsx  (sheet: Phases — only needed if phases missing)

Run:
  python scripts/fixtures/workspace-tasks-4000/gen-workspace-tasks-4000.py
"""

from __future__ import annotations

from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook

OUT_DIR = Path(__file__).resolve().parent

# Exact names from project_phases for DLP & DC Project Plan
PHASES: list[tuple[str, date, date, int]] = [
    ("Phase 1 - Project Initiation & Planning", date(2026, 4, 1), date(2026, 4, 2), 2000),
    ("Phase 2- Design & Documentation", date(2026, 4, 5), date(2026, 4, 29), 4000),
    ("Phase 3- Implementation & UAT", date(2026, 6, 29), date(2026, 9, 13), 12000),
    ("Phase 4- Project Closure", date(2026, 9, 13), date(2026, 9, 16), 2000),
]


def add_days(d: date, n: int) -> date:
    return d + timedelta(days=n)


def clamp(d: date, lo: date, hi: date) -> date:
    return max(lo, min(hi, d))


def task_window(phase_start: date, phase_end: date, task_index: int) -> tuple[date, date]:
    """Keep every task fully inside its phase window."""
    span = max((phase_end - phase_start).days, 0)
    if span == 0:
        return phase_start, phase_end

    # Spread starts across the phase; duration 1–3 days, clamped to phase end
    start = add_days(phase_start, task_index % (span + 1))
    start = clamp(start, phase_start, phase_end)
    wanted_end = add_days(start, 1 + (task_index % 3))
    end = clamp(wanted_end, start, phase_end)
    return start, end


def task_title(phase_no: int, task_index: int) -> str:
    return f"DLP Task {phase_no}.{task_index + 1:04d}"


def main() -> None:
    total_planned = sum(count for *_rest, count in PHASES)
    assert total_planned == 20000, total_planned

    # --- Phases helper (optional — project already has these) ---
    wb_phases = Workbook()
    ws_p = wb_phases.active
    ws_p.title = "Phases"
    ws_p.append(
        ["Name", "Description", "Order", "Status", "Start Date", "End Date"]
    )
    for i, (name, ps, pe, _count) in enumerate(PHASES):
        ws_p.append(
            [
                name,
                f"DLP & DC phase {i + 1}",
                i + 1,
                "Planned",
                ps.isoformat(),
                pe.isoformat(),
            ]
        )
    phases_path = OUT_DIR / "workspace-phases-20000.xlsx"
    wb_phases.save(phases_path)

    # --- Tasks for Import Tasks (Phase names + dates match DLP project) ---
    wb = Workbook(write_only=True)
    ws = wb.create_sheet("Tasks")
    ws.append(
        [
            "Title",
            "Description",
            "Priority",
            "Status",
            "Assignee",
            "Phase",
            "Start Date",
            "End Date",
            "Duration Days",
            "Effort Hours",
            "% Complete",
            "Baseline Start",
            "Baseline End",
            "Baseline Duration Days",
            "Predecessors",
        ]
    )

    deps = 0
    total = 0

    for phase_no, (pname, ps, pe, count) in enumerate(PHASES, start=1):
        last_title: str | None = None
        for t in range(count):
            title = task_title(phase_no, t)
            t_start, t_end = task_window(ps, pe, t)
            duration = max((t_end - t_start).days + 1, 1)

            predecessors = ""
            if t > 0 and (t + 1) % 10 == 1 and last_title:
                predecessors = f"{last_title} (FS)"
                deps += 1

            ws.append(
                [
                    title,
                    f"Load-test task {t + 1} for DLP & DC — {pname}",
                    "High" if (t + 1) % 5 == 0 else "Medium",
                    "To_Do",
                    "",
                    pname,
                    t_start.isoformat(),
                    t_end.isoformat(),
                    str(duration),
                    "16",
                    "0",
                    t_start.isoformat(),
                    t_end.isoformat(),
                    str(duration),
                    predecessors,
                ]
            )
            last_title = title
            total += 1

    tasks_path = OUT_DIR / "workspace-tasks-20000.xlsx"
    wb.save(tasks_path)

    print("wrote", phases_path)
    print("wrote", tasks_path)
    print(f"tasks={total} phases={len(PHASES)} dependencies={deps}")
    print()
    print("Target project: DLP & DC Project Plan (2026-04-01 .. 2026-09-16)")
    print("Phase column must match exactly:")
    for name, ps, pe, count in PHASES:
        print(f"  - {name}  ({ps} .. {pe})  × {count} tasks")


if __name__ == "__main__":
    main()
