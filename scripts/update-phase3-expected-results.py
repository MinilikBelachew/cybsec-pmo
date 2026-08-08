"""Rewrite Expected Result (col L) for Phase "3 - Reporting" rows only.

The register shipped with generated boilerplate ("System successfully executes
the action ... writes them to the 'generated_reports' table") that does not
match what the steps actually verify. Each entry below is derived from the
Test Steps in the same row. No other column or phase is touched.
"""

from __future__ import annotations

import sys
from pathlib import Path

import openpyxl

EXCEL = Path(r"d:\cybsec-pmo\UAT_Test_Case_Register_v2_phase_3.xlsx")
SHEET = "Test Cases"
PHASE = "3 - Reporting"
TC_COL, PHASE_COL, EXPECTED_COL = 1, 2, 12

EXPECTED: dict[str, str] = {
    "TC-M3.1-01": (
        "Dashboard opens with Executive Dashboard / Portfolio Overview as the default tab. "
        "KPI cards render values for projects/progress, at-risk and budget. Execution Health "
        "shows the Project Health table and milestone widgets; People & Resources shows the "
        "utilisation and workload widgets. Settings -> Health rules lists the configured RAG "
        "dimensions (schedule, cost, risk, resources, collections). Budget Burn Rate values "
        "(budget, spent, invoice) remain empty as they are deferred to Phase 5. All dashboard "
        "endpoints return HTTP 200 with no load errors."
    ),
    "TC-M3.1-02": (
        "Access is enforced per role with no data leakage. Engineer: dashboard filters "
        "(Department / Status / Primary PM) are not displayed or applied, and Reports hub paths "
        "requiring reports.view (e.g. /dashboard/reports/utilization) are denied or gated rather "
        "than returning data. PM (own_projects): only projects where the user is Primary or "
        "Secondary PM appear; other PMs' projects are absent. PMO Lead: cross-project visibility "
        "is available with projects from multiple PMs shown."
    ),
    "TC-M3.1-03": (
        "For PMO Lead the filter bar shows Department, Status and Primary PM. Applying any "
        "combination refreshes stats, Project Health, milestones, resources and burn-rate to the "
        "filtered subset, and Clear filters restores the full portfolio. Engineer sees no filter "
        "bar. PM sees Department / Status but not Primary PM, and results stay limited to their "
        "own projects. Each filter change triggers a successful (HTTP 200) refetch."
    ),
    "TC-M3.1-04": (
        "Clicking a row in the Project Health table navigates into the project workspace / "
        "projects list carrying the selected project as context. Returning to People & Resources "
        "shows utilisation KPIs that agree with the detail in Reports -> Utilization "
        "(/dashboard/reports/utilization), which opens successfully for drill-down. No broken "
        "navigation or lost context."
    ),
    "TC-M3.1-05": (
        "A Live / As-of freshness badge is displayed next to Reload. Clicking Reload refetches the "
        "dashboard APIs (HTTP 200) and updates the As-of timestamp. Dashboard endpoints report "
        "their freshness source as live, and no stale values remain after the reload."
    ),
    "TC-M3.1-06": (
        "The report endpoint returns a clean HTTP 404 Not Found with a structured error payload "
        "for the non-existent project ID. The server does not crash, return a 500, or expose "
        "stack traces or schema detail, and remains available for valid requests."
    ),
    "TC-M3.2-01": (
        "Settings -> Health rules lists all five dimensions: schedule, cost, risk, resources and "
        "collections. Evaluating a project (UI health widget or GET /reports/health/projects/{id}) "
        "returns a score and ragStatus (green / amber / red) for every dimension plus an overall "
        "RAG. RAG colours and labels are consistent between the UI and the API response."
    ),
    "TC-M3.2-02": (
        "Super Admin can edit the green / amber / red thresholds for at least two dimensions and "
        "save successfully, with the rules version / active set updating on save. Re-evaluating a "
        "project applies the new thresholds so ragStatus changes in line with the edited "
        "boundaries, and the change persists."
    ),
    "TC-M3.2-03": (
        "Each dimension value reconciles to its source records: task progress and overdue counts, "
        "milestones completed, spend vs budget, open high / critical items, allocation percentage "
        "and collections. After changing a source fact (e.g. completing a milestone or approving "
        "timesheets) and re-evaluating, the affected scores and RAG move in the expected "
        "direction with no stale or contradictory values."
    ),
    "TC-M3.3-01": (
        "The Weekly (WSR) report generates for the selected project and the preview renders a "
        "readable, approved CyberSec layout. The letterhead on every page matches the branding "
        "profile assigned to that project (company name, logo, colours). The control block shows "
        "nine values, three per row: Document reference, Version, Project name, Customer, "
        "Delivered by, Report period, Date issued, Prepared by, Reviewed by. Sections appear in "
        "order: 1. Executive health summary, 2. Milestones, 3. Work completed and work planned, "
        "4. Open action points, 5. Issues, 6. Risks, 7. Pending items, 8. Cost, 9. Missing or "
        "incomplete data, followed by Notes, with each empty section showing a 'nothing to report' "
        "line - Issues and Risks show 'nothing to report' as they are deferred to Phase 5. The "
        "footer carries the document reference and 'Page X of Y', and the Internal Use Only "
        "watermark is present. PDF and DOCX exports download and match the preview."
    ),
    "TC-M3.3-02": (
        "The Monthly (MSR) report generates for the selected project and the preview renders a "
        "readable, approved CyberSec layout. The letterhead on every page matches the branding "
        "profile assigned to that project (company name, logo, colours). The control block shows "
        "nine values, three per row: Document reference, Version, Project name, Customer, "
        "Delivered by, Report period, Date issued, Prepared by, Reviewed by. Sections appear in "
        "order: 1. Executive health summary, 2. Milestones, 3. Work completed and work planned, "
        "4. Open action points, 5. Issues, 6. Risks, 7. Pending items, 8. Cost, 9. Missing or "
        "incomplete data, followed by Notes, with each empty section showing a 'nothing to report' "
        "line - Issues and Risks show 'nothing to report' as they are deferred to Phase 5. The "
        "footer carries the document reference and 'Page X of Y', and the Internal Use Only "
        "watermark is present. PDF and DOCX exports download and match the preview."
    ),
    "TC-M3.3-03": (
        "The report preview (/dashboard/reports/status/{id}) renders all agreed sections: Project "
        "health, Milestones, Open action points and Data quality issues, together with the header "
        "meta (status, overall RAG, period, generated time). The PDF and DOCX exports contain the "
        "same section headings and content, with no agreed section missing."
    ),
    "TC-M3.3-04": (
        "The generated WSR/MSR lists every unresolved Data Quality flag for the project under Data "
        "quality issues, showing flag type, severity and description. After the flags are resolved "
        "on /dashboard/reports/data-quality and the report is regenerated, the missing-data content "
        "clears or reduces accordingly, so the report reflects the current data quality state."
    ),
    "TC-M3.3-05": (
        "A newly generated report is created with status Draft and only the Approve action is "
        "available - Distribute is not offered while Draft. Approve moves the status to Approved "
        "and the badge updates. Distribute then moves the status to Distributed and recipients "
        "receive an email with subject 'WSR - {project}' / 'MSR - {project}', the body notice and "
        "the PDF attachment. No email is sent at the Draft or Approve stages."
    ),
    "TC-M3.3-06": (
        "PDF, DOCX, Excel (.xlsx) and CSV all download successfully for Draft, Approved and "
        "Distributed reports. Each file opens without corruption and contains the health, "
        "milestones, actions and missing-data content, and the Excel export separates content "
        "into discrete sheets (Health, Milestones, Actions, MissingData) where implemented."
    ),
    "TC-M3.3-07": (
        "Regenerating the same project and report type creates a new row with version incremented "
        "by one, while all prior versions remain listed and can still be opened. Filters (project "
        "/ status / type) and pagination return the correct version history, and no earlier "
        "version is overwritten or lost."
    ),
    "TC-M3.4-01": (
        "The meeting is created with the entered title and scheduled datetime, the attendees "
        "selected from the project team (including at least one Engineer), and the agenda items, "
        "decisions and actions with owners. It appears in the Meetings accordion with the latest "
        "meeting on top, and View/Edit confirms attendees, agenda, decisions and actions persisted "
        "exactly as entered."
    ),
    "TC-M3.4-02": (
        "Generate MoM creates a Draft MoM nested under that meeting, labelled 'Minutes vN' with "
        "status Draft. PDF and DOCX download from the MoM row Export menu and render the interim "
        "CyberSec MoM layout containing attendance, agenda, decisions, actions and the "
        "acknowledgement note."
    ),
    "TC-M3.4-03": (
        "Clicking Review moves the Draft MoM to status Reviewed with the badge updated, and "
        "Distribute becomes the next available action while Review is no longer offered."
    ),
    "TC-M3.4-04": (
        "Generating the MoM a second time for the same meeting increments the version to v2, and "
        "both v1 and v2 remain listed nested under the same meeting and can be opened "
        "independently."
    ),
    "TC-M3.4-05": (
        "Distribute on a Reviewed MoM emails all attendees with the MoM PDF attached and an 'Open "
        "Minutes of Meeting' link, and raises an in-app acknowledgement notification for each "
        "attendee that lands on the MoM tab when opened. The PM's Acknowledgements accordion shows "
        "pending recipients as x/y. The Engineer sees no Meetings list - only Distributed MoMs they "
        "were a recipient of, nested under meeting titles - and clicking Acknowledge disables the "
        "button and relabels it Acknowledged. On refresh the PM sees that engineer as Acknowledged "
        "with a timestamp."
    ),
    "TC-M3.5-01": (
        "Status report exports download successfully as PDF, DOCX, Excel and CSV; Reports -> "
        "Utilization exports CSV; and Dashboard Export Portfolio produces xlsx / csv / pdf "
        "according to the user's role permissions. Every file opens with the correct headings and "
        "values, and formats not permitted for the role are not offered."
    ),
    "TC-M3.5-02": (
        "The Create schedule modal requires Project and report type (WSR or MSR) and accepts the "
        "friendly schedule (WSR: weekday + time; MSR: day-of-month + time), the Active toggle and "
        "multi-selected internal roles. Saving creates a schedule card showing project, type, "
        "cron/summary, recipients and the Active/Inactive control. Toggling Active/Inactive "
        "persists after refresh, and Delete removes the schedule only after confirmation."
    ),
    "TC-M3.5-03": (
        "When scheduled delivery fails, the schedule's lastError is populated and visible in the "
        "UI and the job retries per the configured Bull attempt policy rather than failing "
        "silently. audit_logs contains a delivery-failure event (e.g. REPORT_DELIVERY_FAILED) "
        "with the schedule context."
    ),
    "TC-M3.6-01": (
        "Scan rules save with Missing timesheet and Unapproved timesheet set to Include. Scan all "
        "projects completes and raises flags for missing and unapproved timesheets where such data "
        "exists. Filtering by project, flag type and Open returns the matching flags. Resolving a "
        "flag removes it from the Open list and it appears as Resolved when filtered."
    ),
    "TC-M3.6-02": (
        "Scan rules save with Stale integration set to Include, and Scan all projects creates a "
        "STALE_INTEGRATION flag with a clear description identifying the stale integration where "
        "such data exists."
    ),
    "TC-M3.6-03": (
        "Scan rules save with Incomplete project set to Include, and Scan all projects creates "
        "INCOMPLETE_PROJECT flags for the projects with incomplete data where such data exists."
    ),
    "TC-M3.6-04": (
        "The Scan rules page lists each check with Include / Exclude segmented controls and saves "
        "the selection. With one type set to Exclude, a rescan creates or updates no flags of that "
        "type while the included types still appear. Setting it back to Include and rescanning "
        "allows that type to be raised again, confirming the saved rules gate the scan."
    ),
}


def main() -> None:
    sys.stdout.reconfigure(encoding="utf-8")
    wb = openpyxl.load_workbook(EXCEL)
    ws = wb[SHEET]

    updated, skipped = [], []
    seen = set()

    for r in range(2, ws.max_row + 1):
        if str(ws.cell(r, PHASE_COL).value or "").strip() != PHASE:
            continue
        tc = str(ws.cell(r, TC_COL).value or "").strip()
        text = EXPECTED.get(tc)
        if not text:
            skipped.append((r, tc))
            continue
        ws.cell(r, EXPECTED_COL).value = text
        updated.append((r, tc))
        seen.add(tc)

    unused = sorted(set(EXPECTED) - seen)
    if skipped or unused:
        print("NOT UPDATED (no text defined):", skipped)
        print("UNUSED KEYS (no matching row):", unused)
        print("Aborting without saving to avoid a partial edit.")
        sys.exit(1)

    wb.save(EXCEL)
    print(f"Updated Expected Result for {len(updated)} '{PHASE}' rows.")
    for row, tc in updated:
        print(f"  row {row}: {tc}")


if __name__ == "__main__":
    main()
